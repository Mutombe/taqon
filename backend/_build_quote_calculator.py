"""Build the office quote calculator workbook from the live pricing
constants and area-distance data. One-shot generator — run it again to
regenerate after a rate or area change.

  python _build_quote_calculator.py
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Live pricing constants (kept in sync with apps/solar_config/engine/constants.py)
from apps.solar_config.engine.constants import PRICING

SUNDRIES = float(PRICING['sundries_rate'])     # 0.005
LABOUR = float(PRICING['labour_rate'])         # 0.08
PER_KM = float(PRICING['transport_per_km'])    # 0.65

with open(r'C:\Users\PC\documents\taqon\_areas_export.json', encoding='utf-8') as f:
    DATA = json.load(f)
AREAS = DATA['areas']
HQ = DATA['hq']  # [lat, lng]

# ── Styling ──────────────────────────────────────────────────────────
ORANGE = 'F26522'
INK = '0D0D0D'
CREAM = 'FFFBF5'
LIGHT = 'FCEFE6'
GREY = '6B7280'

H1 = Font(name='Calibri', size=18, bold=True, color=INK)
H2 = Font(name='Calibri', size=12, bold=True, color=ORANGE)
LBL = Font(name='Calibri', size=11, color=INK)
MUTED = Font(name='Calibri', size=9, italic=True, color=GREY)
BOLD = Font(name='Calibri', size=11, bold=True, color=INK)
INPUT_FONT = Font(name='Calibri', size=11, bold=True, color='1A56DB')
TOTAL_FONT = Font(name='Calibri', size=14, bold=True, color=ORANGE)
WHITE_BOLD = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

orange_fill = PatternFill('solid', fgColor=ORANGE)
light_fill = PatternFill('solid', fgColor=LIGHT)
input_fill = PatternFill('solid', fgColor='FFF7E6')
cream_fill = PatternFill('solid', fgColor=CREAM)
thin = Side(style='thin', color='E5E0D8')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
USD = '"USD" #,##0.00'
PCT = '0.0%'

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════
# Sheet 1 — Quote Calculator
# ═══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Quote Calculator'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 46

def cell(ref, value=None, font=None, fill=None, fmt=None, align=None, bd=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical='center')
    if bd:
        c.border = border
    return c

cell('B2', 'TAQON ELECTRICO — Quote Calculator', H1)
cell('B3', 'Same engine the website uses. Enter the equipment cost and pick the '
           'installation area; everything else computes automatically.', MUTED)

# RATES
cell('B5', 'RATES  (adjust only if pricing policy changes)', H2)
cell('B6', 'Sundries rate (% of materials)', LBL)
cell('C6', SUNDRIES, INPUT_FONT, input_fill, PCT, 'center', True)
cell('B7', 'Labour rate (% of materials + sundries)', LBL)
cell('C7', LABOUR, INPUT_FONT, input_fill, PCT, 'center', True)
cell('B8', 'Transport rate (USD per km from Harare HQ)', LBL)
cell('C8', PER_KM, INPUT_FONT, input_fill, USD, 'center', True)

# INPUTS
cell('B10', 'INPUTS', H2)
cell('B11', 'Equipment / material cost (USD)', BOLD)
cell('C11', 0, INPUT_FONT, input_fill, USD, 'center', True)
cell('D11', '← enter the sum of all component prices', MUTED)

cell('B12', 'Installation area', BOLD)
cell('C12', AREAS[0]['name'], INPUT_FONT, input_fill, None, 'center', True)
cell('D12', '← pick from the list (drop-down)', MUTED)

cell('B13', 'Distance from HQ (km)', LBL)
cell('C13', "=IFERROR(VLOOKUP(C12,'Area Distances'!$A:$C,3,FALSE),\"\")",
     LBL, light_fill, '0', 'center', True)
cell('D13', 'auto-filled from the selected area', MUTED)

cell('B14', 'Manual distance override (km)', LBL)
cell('C14', None, INPUT_FONT, input_fill, '0', 'center', True)
cell('D14', 'optional — only if the area is not listed (see Custom Distance tab)', MUTED)

cell('B15', 'Distance used', BOLD)
cell('C15', '=IF(C14<>"",C14,C13)', BOLD, light_fill, '0', 'center', True)

# QUOTE
cell('B17', 'QUOTE', H2)
rows = [
    ('Materials (equipment)', '=C11'),
    ('Sundries  (rate × materials)', '=C11*C6'),
    ('Labour  (rate × [materials + sundries])', '=(C11+C19)*C7'),
    ('Transport  (distance × rate/km)', '=C15*C8'),
]
# Note: 'Sundries' lives in row 19 → referenced by Labour above.
r = 18
for label, formula in rows:
    cell(f'B{r}', label, LBL, None, None, None, True)
    cell(f'C{r}', formula, BOLD, cream_fill, USD, 'right', True)
    r += 1

cell('B22', 'TOTAL', WHITE_BOLD, orange_fill, None, 'left', True)
cell('C22', '=C18+C19+C20+C21', TOTAL_FONT, orange_fill, USD, 'right', True)
ws['B22'].fill = orange_fill
ws.row_dimensions[22].height = 24

cell('B24', 'On the customer PDF, Sundries is folded into the Materials line '
            '(Materials shown = Materials + Sundries). The Total is identical.', MUTED)
ws.merge_cells('B24:D24')

# Area drop-down
dv = DataValidation(type='list', formula1="'Area Distances'!$A$2:$A$%d" % (len(AREAS) + 1),
                    allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws['C12'])

# ═══════════════════════════════════════════════════════════════════
# Sheet 2 — Area Distances
# ═══════════════════════════════════════════════════════════════════
ad = wb.create_sheet('Area Distances')
ad.sheet_view.showGridLines = False
ad.column_dimensions['A'].width = 30
ad.column_dimensions['B'].width = 24
ad.column_dimensions['C'].width = 18
for col, title in [('A', 'Area'), ('B', 'Province'), ('C', 'Distance (km from HQ)')]:
    c = ad[f'{col}1']
    c.value = title
    c.font = WHITE_BOLD
    c.fill = PatternFill('solid', fgColor=INK)
    c.alignment = Alignment(horizontal='left', vertical='center')
ad.row_dimensions[1].height = 20
for i, a in enumerate(AREAS, start=2):
    ad[f'A{i}'] = a['name']
    ad[f'B{i}'] = a['province']
    ad[f'C{i}'] = a['distance']
    ad[f'C{i}'].number_format = '0'
    if i % 2 == 0:
        for col in 'ABC':
            ad[f'{col}{i}'].fill = cream_fill
ad.freeze_panes = 'A2'
ad.auto_filter.ref = f'A1:C{len(AREAS) + 1}'

# ═══════════════════════════════════════════════════════════════════
# Sheet 3 — Custom Distance (Haversine)
# ═══════════════════════════════════════════════════════════════════
cd = wb.create_sheet('Custom Distance')
cd.sheet_view.showGridLines = False
cd.column_dimensions['A'].width = 3
cd.column_dimensions['B'].width = 34
cd.column_dimensions['C'].width = 20
cd.column_dimensions['D'].width = 40

def ccell(ref, value=None, font=None, fill=None, fmt=None, align=None, bd=False):
    c = cd[ref]
    if value is not None:
        c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical='center')
    if bd: c.border = border
    return c

ccell('B2', 'Custom Location Distance', H1)
ccell('B3', 'For a site that is not in the Area Distances list. Get the GPS '
            'coordinates from Google Maps (right-click → the lat, lng numbers).', MUTED)
ccell('B5', 'Taqon HQ latitude', LBL); ccell('C5', HQ[0], LBL, light_fill, '0.000000', 'center', True)
ccell('B6', 'Taqon HQ longitude', LBL); ccell('C6', HQ[1], LBL, light_fill, '0.000000', 'center', True)
ccell('B8', 'Site latitude', BOLD); ccell('C8', -17.83, INPUT_FONT, input_fill, '0.000000', 'center', True)
ccell('B9', 'Site longitude', BOLD); ccell('C9', 31.05, INPUT_FONT, input_fill, '0.000000', 'center', True)

hav = ('=6371*2*ASIN(SQRT(SIN(RADIANS(C8-C5)/2)^2'
       '+COS(RADIANS(C5))*COS(RADIANS(C8))*SIN(RADIANS(C9-C6)/2)^2))')
ccell('B11', 'Straight-line distance (km)', BOLD)
ccell('C11', hav, BOLD, cream_fill, '0.0', 'center', True)
ccell('B12', 'Suggested road distance (km) ≈ ×1.3', BOLD)
ccell('C12', '=C11*1.3', TOTAL_FONT, light_fill, '0.0', 'center', True)
ccell('B14', 'Haversine gives the straight-line (as-the-crow-flies) distance. '
             'Real roads are longer, so multiply by ~1.3 for a fair transport '
             'figure, or just use the nearest town in the Area Distances tab.', MUTED)
cd.merge_cells('B14:D16')
cd['B14'].alignment = Alignment(wrap_text=True, vertical='top')

# ═══════════════════════════════════════════════════════════════════
# Sheet 4 — Formula
# ═══════════════════════════════════════════════════════════════════
fm = wb.create_sheet('Formula')
fm.sheet_view.showGridLines = False
fm.column_dimensions['A'].width = 3
fm.column_dimensions['B'].width = 90
lines = [
    ('TAQON ELECTRICO — Pricing Formula', H1),
    ('', None),
    ('The full price is built up in four steps from the equipment cost:', LBL),
    ('', None),
    ('1.  Materials      = sum of all component prices (the equipment cost)', BOLD),
    (f'2.  Sundries       = Materials × {SUNDRIES:.1%}   (consumables)', BOLD),
    (f'3.  Labour         = (Materials + Sundries) × {LABOUR:.0%}', BOLD),
    (f'4.  Transport      = Distance (km from HQ) × USD {PER_KM:.2f}', BOLD),
    ('', None),
    ('     TOTAL         = Materials + Sundries + Labour + Transport', H2),
    ('', None),
    ('Distance:', H2),
    ('  • Known towns/suburbs: use the value in the "Area Distances" tab', LBL),
    ('    (real road distances from Taqon HQ in Strathaven, Harare).', LBL),
    ('  • Anywhere else: use the "Custom Distance" tab — enter the site GPS', LBL),
    ('    coordinates and it computes the distance (haversine × ~1.3 for road).', LBL),
    ('', None),
    ('Worked example (Marondera, 72 km, USD 2,371.20 equipment):', H2),
    ('  Sundries  = 2,371.20 × 0.5%  = 11.86', LBL),
    ('  Labour    = (2,371.20 + 11.86) × 8% = 190.64', LBL),
    ('  Transport = 72 × 0.65 = 46.80', LBL),
    ('  TOTAL     = 2,371.20 + 11.86 + 190.64 + 46.80 = USD 2,620.50', BOLD),
    ('', None),
    ('Rates live in the "Quote Calculator" tab (orange cells) — change them there', MUTED),
    ('and every figure updates. They mirror the website engine as of this export.', MUTED),
]
r = 2
for text, font in lines:
    if text:
        c = fm[f'B{r}']
        c.value = text
        c.font = font or LBL
    r += 1

OUT = r'C:\Users\PC\documents\taqon\Taqon-Quote-Calculator.xlsx'
wb.save(OUT)
print('saved:', OUT)
print(f'rates: sundries={SUNDRIES:.1%} labour={LABOUR:.0%} transport=USD{PER_KM}/km')
print(f'areas: {len(AREAS)}')
