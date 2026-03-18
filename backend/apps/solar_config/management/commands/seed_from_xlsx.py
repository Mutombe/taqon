"""
Seed solar components, package families, and packages from 'packages draft.xlsx'.

Usage:
    python manage.py seed_from_xlsx                          # uses default path
    python manage.py seed_from_xlsx --file /path/to/file.xlsx
    python manage.py seed_from_xlsx --clear                  # clears existing XLSX-seeded data first

The XLSX structure (packages draft.xlsx):
- Row 9: Package names across columns B-AC (28 packages)
- Rows 10-109: Components with names in col A, prices in col AD
- Each cell in columns B-AC has quantity for that component in that package
- Components categorized by row range:
    10-21 = inverter, 22-27 = battery, 28-29 = panel, 30-34 = mounting, 35-109 = accessories

This command is idempotent — uses update_or_create with xlsx_row_key.
"""

import os
import re
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.solar_config.models import (
    SolarComponent, SolarPackageTemplate, PackageComponent, PackageFamily,
)


# Row ranges for component categories (1-indexed, matching XLSX)
CATEGORY_RANGES = [
    (10, 21, 'inverter'),
    (22, 27, 'battery'),
    (28, 29, 'panel'),
    (30, 34, 'mounting'),
    (35, 109, 'accessory'),
]

# Family detection patterns — ordered most-specific first
# Matched against actual XLSX names like "HOME LUXURY 1.0 5KVA", "HOME DELUX 2.0 8KVA",
# "ULTRA POWER V2.0 10KVA", "PREMIUM POWER 1.0 12KVA", "PRO POWER 1.0", "MASTER POWER V1.0"
FAMILY_PATTERNS = [
    (r'master\s*power', '16kva'),
    (r'pro\s*power', '12kva'),
    (r'prem[iu]+m\s*power|prem[iu]+m.*12', '10kva'),      # "PREMUIM/PREMIUM POWER ... 12KVA"
    (r'ultra\s*power.*10|10.*ultra', '8kva'),             # "ULTRA POWER ... 10KVA"
    (r'delux.*8|8.*delux', '5kva_deluxe'),                # "HOME DELUX/DELUXE ... 8KVA"
    (r'(?:home\s*)?luxury.*5|5\s*kva.*luxury', '5kva'),   # "HOME LUXURY ... 5KVA"
    (r'economy.*3|3\s*kva', '3kva'),                      # "Home ECONOMY 3KVA"
]

FAMILY_INFO = {
    '3kva': {'kva_rating': Decimal('3'), 'name': 'Home Economy 3kVA', 'tier': 'starter',
             'description': 'Entry-level solar system for basic household needs.',
             'suitable_for': ['residential', 'small_home']},
    '5kva': {'kva_rating': Decimal('5'), 'name': 'Home Luxury 5kVA', 'tier': 'popular',
             'description': 'Our most popular mid-range system for comfortable, uninterrupted living.',
             'suitable_for': ['residential', 'home_office']},
    '5kva_deluxe': {'kva_rating': Decimal('8'), 'name': 'Home Deluxe 8kVA', 'tier': 'premium',
                    'description': 'Full-house solar solution with expanded battery storage for maximum independence.',
                    'suitable_for': ['residential', 'large_home']},
    '8kva': {'kva_rating': Decimal('10'), 'name': 'Ultra Power 10kVA', 'tier': 'commercial',
             'description': 'High-capacity system for large homes, office complexes, and small institutions.',
             'suitable_for': ['residential', 'small_business', 'office']},
    '10kva': {'kva_rating': Decimal('12'), 'name': 'Premium Power 12kVA', 'tier': 'commercial',
              'description': 'Commercial-grade power with smart monitoring for businesses and institutions.',
              'suitable_for': ['commercial', 'retail', 'institution']},
    '12kva': {'kva_rating': Decimal('16'), 'name': 'ProPower 16kVA', 'tier': 'commercial',
              'description': 'Professional-grade system for large commercial properties.',
              'suitable_for': ['commercial', 'lodge', 'multi_storey']},
    '16kva': {'kva_rating': Decimal('20'), 'name': 'MasterPower 20kVA', 'tier': 'commercial',
              'description': 'Enterprise-grade solar for large estates and commercial buildings.',
              'suitable_for': ['enterprise', 'commercial', 'institution']},
}


def _get_category(row_num):
    """Determine component category from row number."""
    for start, end, category in CATEGORY_RANGES:
        if start <= row_num <= end:
            return category
    return 'accessory'


def _detect_family_code(package_name):
    """Detect which family a package name belongs to."""
    name_lower = package_name.lower()
    for pattern, code in FAMILY_PATTERNS:
        if re.search(pattern, name_lower):
            return code
    return None


def _is_shop_visible(category):
    """Only panels, inverters, and batteries are visible in the shop."""
    return category in ('panel', 'inverter', 'battery')


class Command(BaseCommand):
    help = 'Seed components, families, and packages from packages draft.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', type=str, default=None,
            help='Path to the XLSX file. Defaults to "packages draft.xlsx" in the project root.',
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Delete all XLSX-seeded components and packages before seeding.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required. Install with: pip install openpyxl')

        # Resolve file path
        file_path = options['file']
        if not file_path:
            # Try common locations
            candidates = [
                os.path.join(os.getcwd(), 'packages draft.xlsx'),
                os.path.join(os.path.dirname(os.getcwd()), 'packages draft.xlsx'),
                os.path.join(os.path.expanduser('~'), 'documents', 'taqon', 'packages draft.xlsx'),
            ]
            for candidate in candidates:
                if os.path.exists(candidate):
                    file_path = candidate
                    break

        if not file_path or not os.path.exists(file_path):
            raise CommandError(
                f'XLSX file not found. Please specify with --file or place '
                f'"packages draft.xlsx" in the project root.'
            )

        self.stdout.write(f'Loading XLSX from: {file_path}')

        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb.active

        if options['clear']:
            self._clear_existing()

        # No wrapping transaction — remote DBs may drop connections on long ops
        components = self._seed_components(ws)
        families = self._seed_families()
        packages = self._seed_packages(ws, components, families)

        self.stdout.write(self.style.SUCCESS(
            f'\nDone! Seeded {len(components)} components, '
            f'{len(families)} families, {len(packages)} packages.'
        ))

    def _clear_existing(self):
        """Clear XLSX-seeded data."""
        # Delete packages that have XLSX-seeded components
        pkg_count = SolarPackageTemplate.objects.filter(
            items__component__xlsx_row_key__gt=''
        ).distinct().delete()[0]

        comp_count = SolarComponent.objects.filter(
            xlsx_row_key__gt=''
        ).delete()[0]

        family_count = PackageFamily.objects.all().delete()[0]

        self.stdout.write(self.style.WARNING(
            f'Cleared: {comp_count} components, {pkg_count} packages, {family_count} families'
        ))

    def _seed_components(self, ws):
        """Extract and seed components from column A (names) and AD (prices)."""
        components = {}
        created = 0
        updated = 0

        for row_num in range(10, 110):  # Rows 10-109
            name_cell = ws.cell(row=row_num, column=1)  # Column A
            price_cell = ws.cell(row=row_num, column=30)  # Column AD (30th column)

            name = str(name_cell.value or '').strip()
            if not name:
                continue

            # Get price — handle formulas and raw values
            price_val = price_cell.value
            if price_val is None:
                price_val = 0
            try:
                price = Decimal(str(price_val)).quantize(Decimal('0.01'))
            except Exception:
                price = Decimal('0.00')

            category = _get_category(row_num)
            xlsx_key = f"row_{row_num}_{name[:100]}"

            obj, was_created = SolarComponent.objects.update_or_create(
                xlsx_row_key=xlsx_key,
                defaults={
                    'name': name,
                    'category': category,
                    'price': max(price, Decimal('0')),
                    'shop_visible': _is_shop_visible(category),
                    'is_active': True,
                    'sort_order': row_num * 10,
                },
            )
            components[row_num] = obj

            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(f'Components: {created} created, {updated} updated ({len(components)} total)')
        return components

    def _seed_families(self):
        """Create PackageFamily objects for each family code."""
        families = {}

        for i, (code, info) in enumerate(FAMILY_INFO.items()):
            obj, created = PackageFamily.objects.update_or_create(
                family_code=code,
                defaults={
                    'name': info['name'],
                    'kva_rating': info['kva_rating'],
                    'description': info.get('description', ''),
                    'suitable_for': info.get('suitable_for', []),
                    'is_active': True,
                    'sort_order': i * 10,
                },
            )
            families[code] = obj
            status = 'created' if created else 'updated'
            self.stdout.write(f'  Family: {info["name"]} ({status})')

        return families

    def _seed_packages(self, ws, components, families):
        """Extract and seed packages from row 9 headers and component quantities."""
        packages = []

        # Row 9 has package names across columns B (2) to AC (29) = 28 packages
        for col_num in range(2, 30):
            pkg_name_cell = ws.cell(row=9, column=col_num)
            pkg_name = str(pkg_name_cell.value or '').strip()
            if not pkg_name:
                continue

            # Detect family
            family_code = _detect_family_code(pkg_name)
            family = families.get(family_code) if family_code else None
            family_info = FAMILY_INFO.get(family_code, {})

            # Determine tier from family info
            tier = family_info.get('tier', 'starter')

            # Extract kVA from name (e.g. "HOME LUXURY 1.0 5KVA" → 5.0)
            kva_match = re.search(r'(\d+)\s*kva', pkg_name, re.IGNORECASE)
            inverter_kva = Decimal(str(kva_match.group(1))) if kva_match else family_info.get('kva_rating', Decimal('0'))

            # Extract variant name from the package name
            variant_match = re.search(r'(v?\d+\.?\d*|performance)', pkg_name, re.IGNORECASE)
            variant_name = variant_match.group(0) if variant_match else f'V{col_num - 1}'

            # Create/update the package template
            slug_base = re.sub(r'[^a-z0-9]+', '-', pkg_name.lower()).strip('-')

            pkg, created = SolarPackageTemplate.objects.update_or_create(
                slug=slug_base[:220],
                defaults={
                    'name': pkg_name,
                    'tier': tier,
                    'family': family,
                    'variant_name': variant_name,
                    'inverter_kva': inverter_kva,
                    'suitable_for': family_info.get('suitable_for', []),
                    'is_active': True,
                    'sort_order': col_num * 10,
                },
            )

            # Clear existing items for this package
            pkg.items.all().delete()

            # Add component items
            panel_count = 0
            total_battery_kwh = Decimal('0')
            items_to_create = []

            for row_num in range(10, 110):
                qty_cell = ws.cell(row=row_num, column=col_num)
                qty_val = qty_cell.value
                if qty_val is None or qty_val == 0 or qty_val == '':
                    continue

                try:
                    qty = int(float(str(qty_val)))
                except (ValueError, TypeError):
                    continue

                if qty <= 0:
                    continue

                component = components.get(row_num)
                if not component:
                    continue

                items_to_create.append(
                    PackageComponent(
                        package=pkg,
                        component=component,
                        quantity=qty,
                    )
                )

                # Track panel count and battery capacity
                if component.category == 'panel':
                    panel_count += qty
                elif component.category == 'battery':
                    total_battery_kwh += component.capacity_kwh * qty

            if items_to_create:
                PackageComponent.objects.bulk_create(items_to_create)

            # Update specs
            pkg.panel_count = panel_count
            pkg.battery_capacity_kwh = total_battery_kwh
            pkg.save(update_fields=['panel_count', 'battery_capacity_kwh', 'updated_at'])

            # Recalculate price with dynamic markup
            pkg.recalculate_price()

            status = 'created' if created else 'updated'
            self.stdout.write(
                f'  Package: {pkg_name} ({status}) — '
                f'{len(items_to_create)} components, ${pkg.price}'
            )
            packages.append(pkg)

        return packages
